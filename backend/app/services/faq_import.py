"""Leitura/geração de planilha de FAQ (xlsx/csv) — usado pelo admin e pelo painel."""
import csv
import io


def parse_faq_file(filename: str, content: bytes) -> list[tuple[str, str]]:
    """Lê pares (pergunta, resposta) de um .xlsx (col A/B) ou .csv (2 colunas)."""
    name = (filename or "").lower()
    pares: list[tuple[str, str]] = []

    if name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            q, a = row[0], row[1]
            if q is None or a is None:
                continue
            pares.append((str(q).strip(), str(a).strip()))
    else:  # csv
        text = content.decode("utf-8-sig", errors="replace")
        for row in csv.reader(io.StringIO(text)):
            if len(row) < 2 or not row[0].strip() or not row[1].strip():
                continue
            pares.append((row[0].strip(), row[1].strip()))

    if pares and pares[0][0].lower() in {"pergunta", "question", "perguntas"}:
        pares = pares[1:]
    return pares


def build_template_xlsx() -> bytes:
    """Gera uma planilha modelo com cabeçalho + exemplos."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "FAQ"
    ws.append(["Pergunta", "Resposta"])
    ws.append(["Qual o horário de funcionamento?", "Segunda a sexta, das 6h às 21h."])
    ws.append(["Vocês têm aula experimental?", "Sim! A primeira aula é gratuita."])
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 60
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
